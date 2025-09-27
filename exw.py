from openpyxl import Workbook, load_workbook
from openpyxl.chart import LineChart, Reference
from pathlib import Path

# Arrêt à 677 cases

def create(nom_fichier, titre_feuille, nom_en_tete1, nom_en_tete2):
    fichier = nom_fichier + ".xlsx" # "donnees.xlsx"

    # Création ou ouverture du fichier
    if Path(fichier).exists():
        wb = load_workbook(fichier)
    else:
        wb = Workbook()

    # Vérifier si la feuille existe déjà
    if titre_feuille in wb.sheetnames:
        ws = wb[titre_feuille]
    else:
        ws = wb.create_sheet(title=titre_feuille)
        ws.append([nom_en_tete1, nom_en_tete2])  # En-têtes

    return fichier, ws, wb


# Fonction pour ajouter une donnée et mettre à jour le graphique
def ajouter_donnee(fichier, ws, wb, x, y, titre_graphe, titre1, titre2):
    ws.append([x, y])

    # Supprimer les anciens graphiques (évite doublons)
    for obj in ws._charts:
        ws._charts.remove(obj)

    # Créer un nouveau graphique
    chart = LineChart()
    chart.legend = None
    chart.title = titre_graphe# "Évolution des valeurs"
    chart.x_axis.title = titre1 # "Temps / x"
    chart.y_axis.title = titre2 # "Valeur / y"

    # Définir la plage de données
    data = Reference(ws, min_col=2, min_row=1, max_row=ws.max_row)  # y
    categories = Reference(ws, min_col=1, min_row=2, max_row=ws.max_row)  # x
    chart.add_data(data, titles_from_data=False)
    chart.set_categories(categories)

    for serie in chart.series:
        serie.graphicalProperties.line.solidFill = "FF0000"

    # Position du graphique dans la feuille
    ws.add_chart(chart, "E5")

    # Sauvegarde du fichier
    wb.save(fichier)


